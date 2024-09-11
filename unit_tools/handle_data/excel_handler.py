import openpyxl


def get_excel_data(path):
    """
    从指定路径的 Excel 文件中读取数据，并返回格式化后的字典列表。

    Args:
    - path (str): Excel 文件路径

    Returns:
    - list: 包含读取数据的字典列表
    """
    workbook = openpyxl.load_workbook(path)  # 加载 Excel 文件
    results = []  # 存储最终结果的列表

    for sheet_name in workbook.sheetnames:  # 遍历每个工作表
        sh = workbook[sheet_name]   # 获取工作表对象

        if sh.max_row < 1 or sh.max_column < 1:     # 如果工作表没有数据，则跳过
            continue

        rows = list(sh.rows)    # 获取所有行

        if len(rows) < 1:   # 如果行数小于1，则跳过
            continue

        title = [cell.value for cell in rows[0] if cell.value]      # 获取标题行的内容

        for row in rows[1:]:    # 遍历除了标题行之外的每一行
            if len(row) < len(title):   # 如果行的列数少于标题行的列数，则跳过
                continue

            data = [cell.value for cell in row]     # 获取当前行的数据

            if data[title.index('是否跳过')] == '是':    # 如果“是否跳过”列的值为“是”，则跳过当前行
                continue
            # 构建基本信息字典
            base_info = {
                'api_name': data[title.index('API名称')],
                'url': data[title.index('请求路径')],
                'method': data[title.index('请求类型')],
                'headers': data[title.index('请求头')],
                'model': data[title.index('模块')]
            }
            # 构建测试用例字典
            test_case = {
                'case_name': data[title.index('用例名称')]
            }
            # 如果有校验参数，则将其转为字典格式
            if data[title.index('校验参数')] is not None:
                test_case['validation'] = eval(data[title.index('校验参数')])
            # 如果有上传文件地址，则添加到测试用例字典中
            if data[title.index('上传文件地址')] is not None:
                test_case['files'] = data[title.index('上传文件地址')]

            request_params = data[title.index('请求参数')]
            if request_params is not None:
                request_params = request_params.replace('null', 'None')
            # 根据请求类型区分参数类型
            if base_info.get('method') == 'get':
                test_case['params'] = eval(request_params) if request_params else {}
            else:
                test_case['json'] = eval(request_params) if request_params else {}
            # 构建最终结果字典
            result = {
                'baseInfo': base_info,
                'testCase': [test_case]
            }

            results.append(result)      # 将结果字典添加到结果列表中

    return results      # 返回最终的结果列表
